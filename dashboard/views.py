# dashboard/views.py

from datetime import date, datetime, time
import calendar
from collections import OrderedDict
from branches.models import Branch

import os
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.db.models.functions import TruncMonth, TruncDay
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.contrib.auth.models import Group
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.core.serializers.json import DjangoJSONEncoder
import json
from openpyxl import Workbook 
from members.models import Member, Policy, DiySignupLog
from schemes.models import Plan, Scheme
from settings_app.models import Agent, Underwriter
from reports.models import Report
from config.permissions import filter_by_user_scope, get_user_schemes, get_user_branches


def get_active_policies_queryset(queryset=None):
    if queryset is None:
        queryset = Policy.objects.all()

    today = timezone.now().date()
    return queryset.exclude(lapse_warning='lapsed').filter(
        Q(cover_date__isnull=True) | Q(cover_date__gte=today)
    )


def normalize_activity_date(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, time.min)
    return datetime.min


# ─── MAIN DASHBOARD ───────────────────────────────────────────────────────────
@login_required
def index(request):
    user = request.user
    is_superuser = user.is_superuser

    selected_scheme = request.GET.get('scheme', '')
    selected_month = request.GET.get('month', '')

    # Get user's accessible schemes using multi-tenancy helper
    accessible_schemes = get_user_schemes(user)
    
    # Base querysets with multi-tenancy filtering
    schemes = filter_by_user_scope(Plan.objects.all(), user, Plan)
    members_qs = filter_by_user_scope(Member.objects.all(), user, Member)

    # Filter by scheme if provided
    if selected_scheme.isdigit():
        members_qs = members_qs.filter(
            pk__in=Policy.objects.filter(plan_id=int(selected_scheme))
                                   .values_list('member_id', flat=True)
        )

    # Filter by signup month if provided (based on Policy.inception_date)
    if selected_month.isdigit() and 1 <= int(selected_month) <= 12:
        members_qs = members_qs.filter(
            pk__in=Policy.objects.filter(inception_date__month=int(selected_month))
                                   .values_list('member_id', flat=True)
        )

    total_members = members_qs.count()
    total_schemes = schemes.count()
    this_month = timezone.now().month

    # Policies this month
    monthly_signup = Policy.objects.filter(
        member__in=members_qs,
        inception_date__month=this_month
    ).count()

    total_policies = Policy.objects.filter(member__in=members_qs).count()

    # Monthly aggregation of policies
    monthly_agg = (
        Policy.objects.filter(member__in=members_qs)
        .annotate(month=TruncMonth('inception_date'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    monthly_labels = [entry['month'].strftime('%b') for entry in monthly_agg]
    monthly_data = [entry['count'] for entry in monthly_agg]

    # Plan distribution (number of policies per plan)
    plan_labels = [plan.name for plan in schemes]
    plan_data = [
        Policy.objects.filter(plan=plan, member__in=members_qs).count()
        for plan in schemes
    ]

    # Scheme distribution (number of policies per scheme)
    distinct_schemes = Scheme.objects.filter(plans__in=schemes).distinct()
    scheme_labels = [s.name for s in distinct_schemes]
    scheme_data = [
        Policy.objects.filter(plan__scheme=scheme, member__in=members_qs).count()
        for scheme in distinct_schemes
    ]

    # Branch distribution (number of policies per branch)
    branches = Branch.objects.all()
    branch_names = [b.name for b in branches]
    branch_data = []
    for b in branches:
        branch_data.append(Policy.objects.filter(plan__scheme__branch=b).count())

    # Daily signups for this month
    start_month = timezone.now().replace(day=1)
    agent_signup_data = (
        Policy.objects.filter(inception_date__gte=start_month, member__in=members_qs)
        .annotate(day=TruncDay('inception_date'))
        .values('day')
        .annotate(total=Count('id'))
        .order_by('day')
    )
    agent_signup_labels = [x['day'].strftime('%d %b') for x in agent_signup_data]
    agent_signup_counts = [x['total'] for x in agent_signup_data]

    # Schemes with zero policies this month (for tracking empty schemes)
    empty_schemes = Scheme.objects.annotate(
        mcount=Count(
            'plans__policy',
            filter=Q(plans__policy__inception_date__month=this_month)
        )
    ).filter(mcount=0)

    # Top agents based on the number of policies they've underwritten
    agent_stats = (
        Policy.objects.filter(member__in=members_qs)
        .values('underwritten_by')
        .annotate(total=Count('id'))
        .order_by('-total')[:5]
    )
    agent_ids = [x['underwritten_by'] for x in agent_stats]
    agent_map = {a.id: a for a in Agent.objects.filter(id__in=agent_ids)}
    top_agents = []
    for stat in agent_stats:
        aid = stat['underwritten_by']
        agent = agent_map.get(aid)
        if agent:
            top_agents.append({
                'name': agent.full_name,
                'scheme': agent.scheme.name if agent.scheme else '—',
                'total': stat['total'],
            })

    # Recent members by policy inception date with safe attribute access
    recent_members = []
    recent_policies = Policy.objects.select_related('member', 'plan__scheme__branch').order_by('-inception_date')[:5]
    
    for policy in recent_policies:
        member = policy.member
        if hasattr(member, 'policy'):
            member.branch_name = getattr(policy.plan.scheme.branch, "name", "N/A")
        else:
            member.branch_name = "N/A"
        recent_members.append(member)

    months = [(str(i), calendar.month_name[i]) for i in range(1, 13)]

    context = {
        'schemes': schemes,
        'selected_scheme': selected_scheme,
        'months': months,
        'selected_month': selected_month,
        'total_members': total_members,
        'total_schemes': total_schemes,
        'monthly_signups': monthly_signup,
        'total_policies': total_policies,
        'monthly_labels': monthly_labels,
        'monthly_data': monthly_data,
        'plan_labels': plan_labels,
        'plan_data': plan_data,
        'scheme_labels': scheme_labels,
        'scheme_data': scheme_data,
        'branch_labels': branch_names,
        'branch_data': branch_data,
        'agent_signup_labels': agent_signup_labels,
        'agent_signup_counts': agent_signup_counts,
        'recent_members': recent_members,
        'top_agents': top_agents,
        'inactive_schemes': empty_schemes,
        'is_superuser': is_superuser,
        'is_branch_owner': is_branch_owner,
        'is_scheme_admin': is_scheme_admin,
    }

    # Corrected template path
    template_path = 'dashboard/index.html'
    return render(request, template_path, context)


# ─── BRANCH DASHBOARD ─────────────────────────────────────────────────────────
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Count
from members.models import Policy, Member
from settings_app.models import Agent
from branches.models import Branch
from schemes.models import Scheme

@login_required
def branch_dashboard(request):
    user = request.user

    # Only superuser or Branch Owner allowed
    if not user.is_superuser and not user.groups.filter(name="Branch Owner").exists():
        return render(request, "403.html")

    # Superuser sees all
    if user.is_superuser:
        branch = None
        schemes = Scheme.objects.all()
        policies = Policy.objects.all()
        members = Member.objects.all()
        agents = Agent.objects.all()
    else:
        branch = getattr(user.userprofile, "branch", None)
        if not branch:
            return render(request, "dashboard/branch_dashboard.html", {"error": "No branch assigned."})

        # Filter schemes and related data for the branch
        schemes = Scheme.objects.all() if user.is_superuser else Scheme.objects.filter(branch=branch)
        policies = Policy.objects.filter(plan__scheme__in=schemes)
        members = Member.objects.filter(policy__in=policies)
        agents = Agent.objects.filter(scheme__in=schemes)

    # Ensure there's data to display
    if not schemes.exists() or not policies.exists() or not members.exists() or not agents.exists():
        return render(request, "dashboard/branch_dashboard.html", {"error": "No data available."})

    # Top agents based on the number of policies they've underwritten
    top_agents = (
        policies.values('agent__full_name')  # Accessing the 'full_name' field of Agent
        .annotate(count=Count('id'))
        .order_by('-count')[:5]
    )

    # Correct the `new_signups` line: filter on `inception_date` of the related `Policy` model
    new_signups = Member.objects.filter(
        policies__in=policies,  # Ensure we are filtering based on related Policy objects
        policies__inception_date__month=timezone.now().month  # Now using the `inception_date` from `Policy`
    ).count()

    context = {
        'branch': branch,
        'total_members': members.count(),
        'total_policies': policies.count(),
        'new_signups': new_signups,
        'scheme_count': schemes.count(),
        'agents': agents,
        'top_agents': top_agents,
        'schemes': schemes,
    }

    return HttpResponse(render_to_string("dashboard/branch_dashboard.html", context, request=request))


# ─── SCHEME DASHBOARD ─────────────────────────────────────────────────────────
@login_required
def scheme_dashboard(request, scheme_id):
    user = request.user
    scheme = get_object_or_404(Scheme, pk=scheme_id)
    if not user.is_superuser and scheme.admin_user != user:
        return render(request, "403.html")

    policies    = Policy.objects.filter(plan__scheme=scheme)
    members     = Member.objects.filter(policy__in=policies)
    top_agents  = (
        policies.values('underwritten_by__full_name')
        .annotate(count=Count('id'))
        .order_by('-count')[:5]
    )

    context = {
        'scheme':       scheme,
        'policy_count': policies.count(),
        'member_count': members.count(),
        'new_signups':  members.filter(inception_date__month=timezone.now().month).count(),
        'top_agents':   top_agents,
    }
    template_path = os.path.join('templates', 'dashboard', 'scheme_dashboard.html')
    return render(request, template_path, context)


# ─── EXPORT PDF ───────────────────────────────────────────────────────────────
from django.shortcuts import render
from django.http import HttpResponse
from django.template.loader import render_to_string
from reports.models import Report  # Ensure you import your Report model if needed

@login_required
def export_dashboard_pdf(request):
    return HttpResponse("PDF export is temporarily disabled.")


# ─── ENHANCED ADMIN DASHBOARD ────────────────────────────────────────────────
@login_required
def admin_dashboard(request):
    """Enhanced admin dashboard with visualizations and metrics"""
    # Check if user is admin/superuser
    if not request.user.is_superuser:
        return redirect('dashboard:index')
    
    today = timezone.now().date()
    current_month = today.month
    current_year = today.year
    
    # Summary counts
    total_plans = Plan.objects.count()
    active_plans = Plan.objects.filter(is_active=True).count()
    inactive_plans = total_plans - active_plans
    
    total_policies = Policy.objects.count()
    active_policies = get_active_policies_queryset().count()
    inactive_policies = total_policies - active_policies
    
    total_members = Member.objects.count()
    total_agents = Agent.objects.count()
    total_underwriters = Underwriter.objects.count() if 'Underwriter' in globals() else 0
    
    # Monthly premium data for chart (last 6 months)
    months = []
    premium_data = []
    policy_counts = []
    
    for i in range(5, -1, -1):  # Last 6 months
        # Calculate month and year
        month = current_month - i
        year = current_year
        if month <= 0:
            month += 12
            year -= 1
        
        # Get month name
        month_name = calendar.month_name[month][:3] + f" '{str(year)[2:]}"  # e.g. Jan '23
        months.append(month_name)
        
        # Get premium sum for this month
        month_policies = Policy.objects.filter(
            created_at__year=year,
            created_at__month=month
        )
        monthly_premium = sum(policy.premium_amount or 0 for policy in month_policies)
        premium_data.append(monthly_premium)
        
        # Count policies created this month
        policy_counts.append(month_policies.count())
    
    # Recent activity - new policies, plan changes, etc.
    recent_policies = Policy.objects.select_related('member', 'plan').order_by('-created_at')[:10]
    recent_plans = Plan.objects.select_related('scheme').order_by('-modified')[:10]
    
    # Combine and sort by date
    recent_activities = [
        {'type': 'policy', 'item': policy, 'date': policy.created_at, 'action': 'created'}
        for policy in recent_policies
    ] + [
        {'type': 'plan', 'item': plan, 'date': plan.modified, 'action': 'modified'}
        for plan in recent_plans
    ]
    
    # Sort by date (newest first) and limit to 10
    recent_activities.sort(key=lambda x: normalize_activity_date(x['date']), reverse=True)
    recent_activities = recent_activities[:10]
    
    context = {
        'total_plans': total_plans,
        'active_plans': active_plans,
        'inactive_plans': inactive_plans,
        'total_policies': total_policies,
        'active_policies': active_policies,
        'inactive_policies': inactive_policies,
        'total_members': total_members,
        'total_agents': total_agents,
        'total_underwriters': total_underwriters,
        'months_json': json.dumps(months, cls=DjangoJSONEncoder),
        'premium_data_json': json.dumps(premium_data, cls=DjangoJSONEncoder),
        'policy_counts_json': json.dumps(policy_counts, cls=DjangoJSONEncoder),
        'recent_activities': recent_activities,
    }
    
    return render(request, 'dashboard/admin_dashboard.html', context)


# ─── EXPORT EXCEL ─────────────────────────────────────────────────────────────
@login_required
def export_dashboard_excel(request):
    """Export dashboard data to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard Data"
    
    # Add headers
    headers = ['Category', 'Metric', 'Value']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Get data similar to the dashboard view
    user = request.user
    is_superuser = user.is_superuser
    
    # Add summary data
    row = 2
    
    # Plans data
    total_plans = Plan.objects.count()
    active_plans = Plan.objects.filter(is_active=True).count()
    inactive_plans = total_plans - active_plans
    
    ws.cell(row=row, column=1, value="Plans")
    ws.cell(row=row, column=2, value="Total")
    ws.cell(row=row, column=3, value=total_plans)
    row += 1
    
    ws.cell(row=row, column=1, value="Plans")
    ws.cell(row=row, column=2, value="Active")
    ws.cell(row=row, column=3, value=active_plans)
    row += 1
    
    ws.cell(row=row, column=1, value="Plans")
    ws.cell(row=row, column=2, value="Inactive")
    ws.cell(row=row, column=3, value=inactive_plans)
    row += 1
    
    # Policies data
    total_policies = Policy.objects.count()
    active_policies = get_active_policies_queryset().count()
    inactive_policies = total_policies - active_policies
    
    ws.cell(row=row, column=1, value="Policies")
    ws.cell(row=row, column=2, value="Total")
    ws.cell(row=row, column=3, value=total_policies)
    row += 1
    
    ws.cell(row=row, column=1, value="Policies")
    ws.cell(row=row, column=2, value="Active")
    ws.cell(row=row, column=3, value=active_policies)
    row += 1
    
    ws.cell(row=row, column=1, value="Policies")
    ws.cell(row=row, column=2, value="Inactive")
    ws.cell(row=row, column=3, value=inactive_policies)
    row += 1
    
    # Members data
    total_members = Member.objects.count()
    ws.cell(row=row, column=1, value="Members")
    ws.cell(row=row, column=2, value="Total")
    ws.cell(row=row, column=3, value=total_members)
    
    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=dashboard_data.xlsx'
    wb.save(response)
    return response


# ─── AGENT DIY DASHBOARD ─────────────────────────────────────────────────────
@login_required
def dashboard(request):
    total_policies   = Policy.objects.count()
    diy_signups_count= DiySignupLog.objects.filter(agent=request.user.agent).count()
    pending_otp_count= Policy.objects.filter(otp_sent=True, otp_verified=False).count()
    return render(request, 'dashboard.html', {
        'total_policies':     total_policies,
        'diy_signups_count':  diy_signups_count,
        'pending_otp_count':  pending_otp_count,
    })
