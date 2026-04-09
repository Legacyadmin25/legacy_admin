import io
from datetime import date
from decimal import Decimal

from django.contrib.auth.models import Group
from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from audit.models import AuditLog
from import_data.models import PolicyAmendmentImport, PolicyAmendmentRowLog
from members.models import Member, Policy
from payments.models import Payment
from schemes.models import Plan, Scheme
from branches.models import Bank, Branch
from settings_app.models import Agent
from django.core.files.uploadedfile import SimpleUploadedFile


class PaymentAllocationReportTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username='reporter',
            password='pass1234',
            is_superuser=True,
            is_staff=True,
        )
        self.client.force_login(self.user)

        self.bank = Bank.objects.create(name='Bank', branch_code='123456')
        self.branch = Branch.objects.create(name='Main Branch', bank=self.bank, code='MB01')
        self.scheme = Scheme.objects.create(
            branch=self.branch,
            name='Test Scheme',
            prefix='TS',
            registration_no='REG1',
            fsp_number='FSP1',
            email='scheme@example.com',
            phone='0123456789',
            debit_order_no='DEB1',
            account_no='12345',
        )
        self.plan = Plan.objects.create(
            name='Gold Plan',
            scheme=self.scheme,
            premium='150.00',
            policy_type='cash',
            main_cover='10000.00',
            main_uw_cover='9000.00',
            main_uw_premium='80.00',
            admin_fee='20.00',
            scheme_fee='10.00',
            office_fee='5.00',
            manager_fee='3.00',
            agent_commission='12.00',
        )
        self.member = Member.objects.create(
            first_name='Jane',
            last_name='Doe',
            gender='Female',
            date_of_birth=date(1990, 1, 1),
            phone_number='0820000000',
        )
        self.agent = Agent.objects.create(
            full_name='Agent Smith',
            surname='Smith',
            contact_number='0821111111',
            email='agent@example.com',
            address1='1 Main',
            address2='Town',
            address3='Province',
            code='AG01',
            scheme=self.scheme,
            commission_rand_value='15.00',
        )
        self.policy = Policy.objects.create(
            member=self.member,
            scheme=self.scheme,
            plan=self.plan,
            membership_number='M123',
            start_date=date(2026, 4, 1),
            cover_date=date(2026, 5, 1),
            inception_date=date(2026, 4, 1),
            payment_method='EASYPAY',
            underwritten_by=self.agent,
            cover_amount='10000.00',
            premium_amount='150.00',
        )

    def test_payment_capture_creates_default_allocation(self):
        payment = Payment.objects.create(
            member=self.member,
            policy=self.policy,
            amount='150.00',
            date=date(2026, 4, 15),
            payment_method='EASYPAY',
            status='COMPLETED',
            created_by=self.user,
        )

        allocation = payment.create_default_allocation(date(2026, 5, 9), created_by=self.user)

        self.assertIsNotNone(allocation)
        self.assertEqual(allocation.coverage_month, date(2026, 5, 1))
        self.assertEqual(allocation.agent_name, 'Agent Smith')
        self.assertEqual(allocation.agent_code, 'AG01')
        self.assertEqual(allocation.retail_premium, Decimal('150.00'))
        self.assertEqual(allocation.agent_commission, Decimal('15.00'))

    def test_payment_allocation_report_uses_allocated_payments(self):
        payment = Payment.objects.create(
            member=self.member,
            policy=self.policy,
            amount='150.00',
            date=date(2026, 4, 15),
            payment_method='EASYPAY',
            status='COMPLETED',
            created_by=self.user,
        )
        payment.create_default_allocation(date(2026, 5, 1), created_by=self.user)

        response = self.client.get(reverse('reports:payment_allocation_report'), {
            'scheme': self.scheme.id,
            'cover_month': '2026-05',
            'version': 'scheme',
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Jane Doe')
        self.assertContains(response, 'Agent Smith')
        self.assertContains(response, '2026-05')

    def test_payment_allocation_report_admin_and_scheme_versions_show_different_columns(self):
        payment = Payment.objects.create(
            member=self.member,
            policy=self.policy,
            amount='150.00',
            date=date(2026, 4, 15),
            payment_method='EASYPAY',
            status='COMPLETED',
            created_by=self.user,
        )
        payment.create_default_allocation(date(2026, 5, 1), created_by=self.user)

        admin_response = self.client.get(reverse('reports:payment_allocation_report'), {
            'scheme': self.scheme.id,
            'cover_month': '2026-05',
            'version': 'admin',
        })
        scheme_response = self.client.get(reverse('reports:payment_allocation_report'), {
            'scheme': self.scheme.id,
            'cover_month': '2026-05',
            'version': 'scheme',
        })

        self.assertContains(admin_response, 'UW Premium')
        self.assertContains(admin_response, 'Admin Fee')
        self.assertNotContains(admin_response, 'Wholesale Amount')
        self.assertContains(scheme_response, 'Wholesale Amount')
        self.assertNotContains(scheme_response, 'UW Premium')

    def test_payment_allocation_report_exports_excel(self):
        payment = Payment.objects.create(
            member=self.member,
            policy=self.policy,
            amount='150.00',
            date=date(2026, 4, 15),
            payment_method='EASYPAY',
            status='COMPLETED',
            created_by=self.user,
        )
        payment.create_default_allocation(date(2026, 5, 1), created_by=self.user)

        response = self.client.get(reverse('reports:payment_allocation_report'), {
            'scheme': self.scheme.id,
            'cover_month': '2026-05',
            'version': 'admin',
            'export': 'excel',
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        workbook = load_workbook(io.BytesIO(response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet['A1'].value, 'Payment Allocation Report May 2026 (Admin)')
        self.assertEqual(worksheet['A4'].value, '2026-04-15')

    def test_all_members_report_exports_excel(self):
        response = self.client.get(reverse('reports:all_members_report'), {
            'scheme': self.scheme.id,
            'export': 'excel',
        })

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.content))
        worksheet = workbook.active
        self.assertEqual(worksheet['A1'].value, 'All Members Report')
        self.assertEqual(worksheet['A4'].value, 'Test Scheme')

    def test_all_members_report_lists_policy_and_member_details(self):
        response = self.client.get(reverse('reports:all_members_report'), {
            'scheme': self.scheme.id,
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.policy.policy_number)
        self.assertContains(response, 'Jane Doe')
        self.assertContains(response, 'Agent Smith')

    def test_all_members_report_filters_by_agent_and_search(self):
        response = self.client.get(reverse('reports:all_members_report'), {
            'branch': self.branch.id,
            'scheme': self.scheme.id,
            'agent': self.agent.id,
            'search': 'Jane',
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Jane Doe')
        self.assertContains(response, 'Agent Smith')

    def test_all_members_report_exports_csv(self):
        response = self.client.get(reverse('reports:all_members_report'), {
            'scheme': self.scheme.id,
            'export': 'csv',
        })
        self.assertEqual(response.status_code, 200)
        self.assertIn('text/csv', response['Content-Type'])
        self.assertIn('Test Scheme', response.content.decode())

    def test_payment_allocation_report_exports_pdf(self):
        payment = Payment.objects.create(
            member=self.member,
            policy=self.policy,
            amount='150.00',
            date=date(2026, 4, 15),
            payment_method='EASYPAY',
            status='COMPLETED',
            created_by=self.user,
        )
        payment.create_default_allocation(date(2026, 5, 1), created_by=self.user)

        response = self.client.get(reverse('reports:payment_allocation_report'), {
            'scheme': self.scheme.id,
            'cover_month': '2026-05',
            'version': 'scheme',
            'export': 'pdf',
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')

    def test_payment_allocation_report_filters_by_branch_agent_and_search(self):
        payment = Payment.objects.create(
            member=self.member,
            policy=self.policy,
            amount='150.00',
            date=date(2026, 4, 15),
            payment_method='EASYPAY',
            status='COMPLETED',
            created_by=self.user,
        )
        payment.create_default_allocation(date(2026, 5, 1), created_by=self.user)

        response = self.client.get(reverse('reports:payment_allocation_report'), {
            'branch': self.branch.id,
            'scheme': self.scheme.id,
            'agent': self.agent.id,
            'search': 'Jane',
            'cover_month': '2026-05',
            'version': 'admin',
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Jane Doe')
        self.assertContains(response, 'Agent Smith')

    def test_amendments_report_includes_import_changes(self):
        batch = PolicyAmendmentImport.objects.create(
            uploaded_by=self.user,
            file=SimpleUploadedFile('amendments.csv', b'membership_number,premium_amount\n'),
            status='completed',
        )
        PolicyAmendmentRowLog.objects.create(
            import_batch=batch,
            row_number=1,
            membership_number=self.policy.membership_number,
            status='success',
            changes={'premium_amount': ['150.00', '175.00']},
        )

        response = self.client.get(reverse('reports:amendments_report'), {
            'scheme': self.scheme.id,
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'premium_amount')
        self.assertContains(response, '175.00')

    def test_amendments_report_includes_audit_changes(self):
        AuditLog.objects.create(
            user=self.user,
            username=self.user.username,
            action='update',
            content_type_id=1,
            object_id=str(self.policy.id),
            object_repr=str(self.policy),
            data={'field': 'policy_type', 'old_value': 'cash', 'new_value': 'service'},
        )
        from django.contrib.contenttypes.models import ContentType
        log = AuditLog.objects.latest('id')
        log.content_type = ContentType.objects.get_for_model(Policy)
        log.save(update_fields=['content_type'])

        response = self.client.get(reverse('reports:amendments_report'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'policy_type')
        self.assertContains(response, 'service')

    def test_amendments_report_exports_csv(self):
        batch = PolicyAmendmentImport.objects.create(
            uploaded_by=self.user,
            file=SimpleUploadedFile('amendments.csv', b'membership_number,premium_amount\n'),
            status='completed',
        )
        PolicyAmendmentRowLog.objects.create(
            import_batch=batch,
            row_number=1,
            membership_number=self.policy.membership_number,
            status='success',
            changes={'premium_amount': ['150.00', '175.00']},
        )

        response = self.client.get(reverse('reports:amendments_report'), {
            'scheme': self.scheme.id,
            'export': 'csv',
        })

        self.assertEqual(response.status_code, 200)
        self.assertIn('text/csv', response['Content-Type'])
        self.assertIn('premium_amount', response.content.decode())

    def test_amendments_report_filters_by_branch_and_search(self):
        batch = PolicyAmendmentImport.objects.create(
            uploaded_by=self.user,
            file=SimpleUploadedFile('amendments.csv', b'membership_number,premium_amount\n'),
            status='completed',
        )
        PolicyAmendmentRowLog.objects.create(
            import_batch=batch,
            row_number=1,
            membership_number=self.policy.membership_number,
            status='success',
            changes={'premium_amount': ['150.00', '175.00']},
        )

        response = self.client.get(reverse('reports:amendments_report'), {
            'branch': self.branch.id,
            'scheme': self.scheme.id,
            'search': 'Jane',
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'premium_amount')
        self.assertContains(response, 'Jane Doe')

    def test_ai_reports_route_is_not_available(self):
        response = self.client.get('/reports/ai/')
        self.assertEqual(response.status_code, 404)

    def test_report_hub_loads_for_superuser(self):
        response = self.client.get(reverse('reports:report_hub'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Generate Report')
        self.assertContains(response, 'Filter Report')
        self.assertContains(response, 'All Members Report')
        self.assertContains(response, 'Payment Report: Admin')
        self.assertContains(response, 'Payment Report: Scheme')
        self.assertContains(response, 'Amendments Report')

    def test_report_hub_redirects_selected_report_with_filters(self):
        response = self.client.get(reverse('reports:report_hub'), {
            'report': 'payment_admin',
            'scheme': self.scheme.id,
            'agent': self.agent.id,
            'search': 'Jane',
            'cover_month': '2026-05',
            'generate': '1',
        })

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('reports:payment_allocation_report'), response['Location'])
        self.assertIn('version=admin', response['Location'])
        self.assertIn(f'scheme={self.scheme.id}', response['Location'])
        self.assertIn(f'agent={self.agent.id}', response['Location'])

    def test_legacy_report_routes_are_removed(self):
        self.assertEqual(self.client.get('/reports/full_policy_report/').status_code, 404)
        self.assertEqual(self.client.get('/reports/plan_fee_report/').status_code, 404)


class ReportPermissionTests(TestCase):
    def setUp(self):
        self.bank = Bank.objects.create(name='Bank', branch_code='123456')
        self.branch = Branch.objects.create(name='Main Branch', bank=self.bank, code='MB01')
        self.scheme = Scheme.objects.create(
            branch=self.branch,
            name='Test Scheme',
            prefix='TS',
            registration_no='REG1',
            fsp_number='FSP1',
            email='scheme@example.com',
            phone='0123456789',
            debit_order_no='DEB1',
            account_no='12345',
        )
        self.scheme_manager = get_user_model().objects.create_user(username='scheme_manager', password='pass1234')
        self.branch_owner = get_user_model().objects.create_user(username='branch_owner', password='pass1234', branch=self.branch)
        self.administrator = get_user_model().objects.create_user(username='administrator', password='pass1234')
        self.scheme_manager.assigned_schemes.add(self.scheme)
        scheme_group, _ = Group.objects.get_or_create(name='Scheme Manager')
        branch_group, _ = Group.objects.get_or_create(name='Branch Owner')
        admin_group, _ = Group.objects.get_or_create(name='Administrator')
        self.scheme_manager.groups.add(scheme_group)
        self.branch_owner.groups.add(branch_group)
        self.administrator.groups.add(admin_group)

    def test_scheme_manager_can_view_scheme_payment_report_only(self):
        self.client.force_login(self.scheme_manager)

        scheme_response = self.client.get(reverse('reports:payment_allocation_report'), {'version': 'scheme'})
        admin_response = self.client.get(reverse('reports:payment_allocation_report'), {'version': 'admin'})
        amendments_response = self.client.get(reverse('reports:amendments_report'))

        self.assertEqual(scheme_response.status_code, 200)
        self.assertEqual(admin_response.status_code, 403)
        self.assertEqual(amendments_response.status_code, 403)

        hub_response = self.client.get(reverse('reports:report_hub'))
        all_members_response = self.client.get(reverse('reports:all_members_report'))
        self.assertContains(all_members_response, reverse('reports:report_hub'))
        self.assertNotContains(all_members_response, 'Payment Report: Admin')
        self.assertNotContains(all_members_response, 'Payment Report: Scheme')
        self.assertNotContains(all_members_response, 'Amendments Report')
        self.assertContains(hub_response, 'Payment Report: Scheme')
        self.assertContains(hub_response, 'All Members Report')
        self.assertNotContains(hub_response, 'Payment Report: Admin')
        self.assertNotContains(hub_response, 'Amendments Report')

    def test_branch_owner_can_view_admin_and_amendments_reports(self):
        self.client.force_login(self.branch_owner)

        hub_response = self.client.get(reverse('reports:report_hub'))
        all_members_response = self.client.get(reverse('reports:all_members_report'))
        admin_response = self.client.get(reverse('reports:payment_allocation_report'), {'version': 'admin'})
        scheme_response = self.client.get(reverse('reports:payment_allocation_report'), {'version': 'scheme'})
        amendments_response = self.client.get(reverse('reports:amendments_report'))

        self.assertEqual(all_members_response.status_code, 200)
        self.assertEqual(admin_response.status_code, 200)
        self.assertEqual(scheme_response.status_code, 200)
        self.assertEqual(amendments_response.status_code, 200)
        self.assertContains(all_members_response, reverse('reports:report_hub'))
        self.assertNotContains(all_members_response, 'Payment Report: Admin')
        self.assertNotContains(all_members_response, 'Payment Report: Scheme')
        self.assertNotContains(all_members_response, 'Amendments Report')
        self.assertContains(hub_response, 'Payment Report: Admin')
        self.assertContains(hub_response, 'Payment Report: Scheme')
        self.assertContains(hub_response, 'Amendments Report')