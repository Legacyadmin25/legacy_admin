from datetime import datetime
from decimal import Decimal
import io
import csv
from urllib.parse import urlencode

from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db.models import Count, DecimalField, ExpressionWrapper, F, Q, Sum
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from django.contrib.contenttypes.models import ContentType
from django.utils.html import escape
from weasyprint import HTML

from config.permissions import (
    can_view_all_members_report,
    can_view_amendments_report,
    can_view_payment_allocation_report,
    filter_by_user_scope,
    get_user_branches,
    get_user_schemes,
    user_has_role,
)
from branches.models import Branch
from members.models import Policy
from payments.models import PaymentAllocation
from schemes.models import Scheme
from import_data.models import PolicyAmendmentRowLog
from audit.models import AuditLog
from settings_app.models import Agent


def parse_month_param(value):
    if not value:
        return None
    for fmt in ('%Y-%m', '%Y-%m-%d'):
        try:
            return datetime.strptime(value, fmt).date().replace(day=1)
        except ValueError:
            continue
    return None


def apply_policy_filters(queryset, filter_scheme, filter_status):
    if filter_scheme:
        queryset = queryset.filter(scheme_id=filter_scheme)

    today = timezone.now().date()
    if filter_status == 'active':
        queryset = queryset.filter(cover_date__gte=today)
    elif filter_status == 'lapsed':
        queryset = queryset.filter(cover_date__lt=today)

    return queryset


def build_report_workbook(title, headers, rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title[:31]

    title_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    header_fill = PatternFill(fill_type='solid', fgColor='D9EAF7')
    bold_font = Font(bold=True, color='FFFFFF')
    header_font = Font(bold=True)

    worksheet.cell(row=1, column=1, value=title)
    worksheet.cell(row=1, column=1).fill = title_fill
    worksheet.cell(row=1, column=1).font = bold_font

    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=3, column=column, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_index, row_data in enumerate(rows, start=4):
        for column, value in enumerate(row_data, start=1):
            worksheet.cell(row=row_index, column=column, value=value)

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 28)

    worksheet.freeze_panes = 'A4'
    return workbook


def workbook_response(workbook, filename):
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def csv_response(headers, rows, filename):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        writer = csv.writer(response)
        writer.writerow(headers)
        for row in rows:
                writer.writerow(row)
        return response


def pdf_response(title, headers, rows, filename):
        table_headers = ''.join(f'<th>{escape(str(header))}</th>' for header in headers)
        table_rows = ''.join(
                '<tr>' + ''.join(f'<td>{escape(str(value))}</td>' for value in row) + '</tr>'
                for row in rows
        )
        html = f"""
        <html>
            <head>
                <style>
                    body {{ font-family: Arial, sans-serif; font-size: 10px; }}
                    h1 {{ font-size: 18px; margin-bottom: 12px; }}
                    table {{ width: 100%; border-collapse: collapse; }}
                    th, td {{ border: 1px solid #cbd5e1; padding: 6px; text-align: left; }}
                    th {{ background: #d9eaf7; }}
                </style>
            </head>
            <body>
                <h1>{escape(title)}</h1>
                <table>
                    <thead><tr>{table_headers}</tr></thead>
                    <tbody>{table_rows}</tbody>
                </table>
            </body>
        </html>
        """
        pdf_bytes = HTML(string=html).write_pdf()
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


def export_all_members_report(policies, cover_title):
    headers = [
        'Scheme', 'Branch', 'Policy Number', 'Membership Number', 'Client Name', 'ID Number', 'Phone',
        'Plan', 'Product Type', 'Retail Rate', 'Cover Amount', 'Underwriter Premium', 'Underwriter Cover',
        'Payment Method', 'Agent Name', 'Agent Code', 'Dependants', 'Beneficiaries', 'Start Date', 'Cover Date'
    ]
    rows = [
        [
            policy.scheme.name if policy.scheme else '',
            policy.scheme.branch.name if policy.scheme and policy.scheme.branch_id else '',
            policy.policy_number or policy.unique_policy_number or '',
            policy.membership_number or '',
            f"{policy.member.first_name} {policy.member.last_name}",
            policy.member.id_number or '',
            policy.member.phone_number or '',
            policy.plan.name if policy.plan else '',
            policy.plan.get_policy_type_display() if policy.plan else '',
            float(policy.plan.premium if policy.plan else policy.premium_amount or 0),
            float(policy.plan.main_cover if policy.plan else policy.cover_amount or 0),
            float(policy.plan.main_uw_premium if policy.plan else 0),
            float(policy.plan.main_uw_cover if policy.plan else 0),
            policy.get_payment_method_display() if policy.payment_method else '',
            policy.underwritten_by.full_name if policy.underwritten_by else '',
            policy.underwritten_by.code if policy.underwritten_by else '',
            policy.dependents_count,
            policy.beneficiaries_count,
            policy.start_date.isoformat() if policy.start_date else '',
            policy.cover_date.isoformat() if policy.cover_date else '',
        ]
        for policy in policies
    ]
    workbook = build_report_workbook(cover_title, headers, rows)
    return workbook_response(workbook, 'all-members-report.xlsx')


def build_all_members_dataset(policies, title='All Members Report'):
    headers = [
        'Scheme', 'Branch', 'Policy Number', 'Membership Number', 'Client Name', 'ID Number', 'Phone',
        'Plan', 'Product Type', 'Retail Rate', 'Cover Amount', 'Underwriter Premium', 'Underwriter Cover',
        'Payment Method', 'Agent Name', 'Agent Code', 'Dependants', 'Beneficiaries', 'Start Date', 'Cover Date'
    ]
    rows = [
        [
            policy.scheme.name if policy.scheme else '',
            policy.scheme.branch.name if policy.scheme and policy.scheme.branch_id else '',
            policy.policy_number or policy.unique_policy_number or '',
            policy.membership_number or '',
            f"{policy.member.first_name} {policy.member.last_name}",
            policy.member.id_number or '',
            policy.member.phone_number or '',
            policy.plan.name if policy.plan else '',
            policy.plan.get_policy_type_display() if policy.plan else '',
            float(policy.plan.premium if policy.plan else policy.premium_amount or 0),
            float(policy.plan.main_cover if policy.plan else policy.cover_amount or 0),
            float(policy.plan.main_uw_premium if policy.plan else 0),
            float(policy.plan.main_uw_cover if policy.plan else 0),
            policy.get_payment_method_display() if policy.payment_method else '',
            policy.underwritten_by.full_name if policy.underwritten_by else '',
            policy.underwritten_by.code if policy.underwritten_by else '',
            policy.dependents_count,
            policy.beneficiaries_count,
            policy.start_date.isoformat() if policy.start_date else '',
            policy.cover_date.isoformat() if policy.cover_date else '',
        ]
        for policy in policies
    ]
    return title, headers, rows


def export_payment_allocation_report(allocations, cover_month, report_version):
    if report_version == 'scheme':
        headers = [
            'Paid Date', 'Cover Month', 'Scheme', 'Policy Number', 'Client Name', 'Product', 'Payment Method',
            'Allocated Amount', 'Retail Rate', 'Wholesale Amount', 'Agent Commission', 'Agent Name', 'Agent Code'
        ]
        rows = [
            [
                allocation.payment.date.isoformat(),
                allocation.coverage_month.strftime('%Y-%m'),
                allocation.scheme.name if allocation.scheme else '',
                allocation.policy.policy_number,
                f"{allocation.member.first_name} {allocation.member.last_name}",
                allocation.product_name or (allocation.plan.name if allocation.plan else ''),
                allocation.payment.get_payment_method_display(),
                float(allocation.allocated_amount),
                float(allocation.retail_premium),
                float(allocation.wholesale_amount),
                float(allocation.agent_commission),
                allocation.agent_name,
                allocation.agent_code,
            ]
            for allocation in allocations
        ]
    else:
        headers = [
            'Paid Date', 'Cover Month', 'Scheme', 'Policy Number', 'Client Name', 'Product', 'Payment Method',
            'Allocated Amount', 'Retail Rate', 'Underwriter Premium', 'Admin Fee', 'Scheme Fee', 'Branch Fee',
            'Manager Fee', 'Cash Payout', 'Loyalty Programme', 'Other Fees', 'Agent Commission', 'Agent Name', 'Agent Code'
        ]
        rows = [
            [
                allocation.payment.date.isoformat(),
                allocation.coverage_month.strftime('%Y-%m'),
                allocation.scheme.name if allocation.scheme else '',
                allocation.policy.policy_number,
                f"{allocation.member.first_name} {allocation.member.last_name}",
                allocation.product_name or (allocation.plan.name if allocation.plan else ''),
                allocation.payment.get_payment_method_display(),
                float(allocation.allocated_amount),
                float(allocation.retail_premium),
                float(allocation.underwriter_premium),
                float(allocation.admin_fee),
                float(allocation.scheme_fee),
                float(allocation.branch_fee),
                float(allocation.manager_fee),
                float(allocation.cash_payout),
                float(allocation.loyalty_programme),
                float(allocation.other_fees),
                float(allocation.agent_commission),
                allocation.agent_name,
                allocation.agent_code,
            ]
            for allocation in allocations
        ]

    workbook = build_report_workbook(
        f"Payment Allocation Report {cover_month.strftime('%B %Y')} ({report_version.title()})",
        headers,
        rows,
    )
    return workbook_response(workbook, f'payment-allocation-{cover_month.strftime("%Y-%m")}-{report_version}.xlsx')


def build_payment_allocation_dataset(allocations, cover_month, report_version):
    title = f"Payment Allocation Report {cover_month.strftime('%B %Y')} ({report_version.title()})"
    if report_version == 'scheme':
        headers = [
            'Paid Date', 'Cover Month', 'Scheme', 'Policy Number', 'Client Name', 'Product', 'Payment Method',
            'Allocated Amount', 'Retail Rate', 'Wholesale Amount', 'Agent Commission', 'Agent Name', 'Agent Code'
        ]
        rows = [
            [
                allocation.payment.date.isoformat(),
                allocation.coverage_month.strftime('%Y-%m'),
                allocation.scheme.name if allocation.scheme else '',
                allocation.policy.policy_number,
                f"{allocation.member.first_name} {allocation.member.last_name}",
                allocation.product_name or (allocation.plan.name if allocation.plan else ''),
                allocation.payment.get_payment_method_display(),
                float(allocation.allocated_amount),
                float(allocation.retail_premium),
                float(allocation.wholesale_amount),
                float(allocation.agent_commission),
                allocation.agent_name,
                allocation.agent_code,
            ] for allocation in allocations
        ]
    else:
        headers = [
            'Paid Date', 'Cover Month', 'Scheme', 'Policy Number', 'Client Name', 'Product', 'Payment Method',
            'Allocated Amount', 'Retail Rate', 'Underwriter Premium', 'Admin Fee', 'Scheme Fee', 'Branch Fee',
            'Manager Fee', 'Cash Payout', 'Loyalty Programme', 'Other Fees', 'Agent Commission', 'Agent Name', 'Agent Code'
        ]
        rows = [
            [
                allocation.payment.date.isoformat(),
                allocation.coverage_month.strftime('%Y-%m'),
                allocation.scheme.name if allocation.scheme else '',
                allocation.policy.policy_number,
                f"{allocation.member.first_name} {allocation.member.last_name}",
                allocation.product_name or (allocation.plan.name if allocation.plan else ''),
                allocation.payment.get_payment_method_display(),
                float(allocation.allocated_amount),
                float(allocation.retail_premium),
                float(allocation.underwriter_premium),
                float(allocation.admin_fee),
                float(allocation.scheme_fee),
                float(allocation.branch_fee),
                float(allocation.manager_fee),
                float(allocation.cash_payout),
                float(allocation.loyalty_programme),
                float(allocation.other_fees),
                float(allocation.agent_commission),
                allocation.agent_name,
                allocation.agent_code,
            ] for allocation in allocations
        ]
    return title, headers, rows


def build_amendments_dataset(entries):
    headers = [
        'Date', 'Source', 'Scheme', 'Policy Number', 'Membership Number', 'Client', 'Entity',
        'Field', 'Old Value', 'New Value', 'Changed By', 'Reference', 'Notes'
    ]
    rows = [
        [
            entry['date'], entry['source'], entry['scheme'], entry['policy_number'], entry['membership_number'],
            entry['client'], entry['entity'], entry['field'], entry['old_value'], entry['new_value'],
            entry['changed_by'], entry['reference'], entry['notes']
        ] for entry in entries
    ]
    return 'Amendments Report', headers, rows


def collect_amendment_entries(filter_scheme='', filter_branch='', start_date='', end_date='', source='', search_term=''):
    entries = []

    import_logs = PolicyAmendmentRowLog.objects.filter(status='success').exclude(changes={}).select_related(
        'import_batch', 'import_batch__uploaded_by'
    ).order_by('-import_batch__uploaded_at', '-row_number')
    if start_date:
        import_logs = import_logs.filter(import_batch__uploaded_at__date__gte=start_date)
    if end_date:
        import_logs = import_logs.filter(import_batch__uploaded_at__date__lte=end_date)

    for log in import_logs:
        policy = Policy.objects.select_related('member', 'scheme').filter(membership_number=log.membership_number).first()
        if filter_scheme and (not policy or str(policy.scheme_id) != str(filter_scheme)):
            continue
        if filter_branch and (not policy or str(policy.scheme.branch_id if policy and policy.scheme_id else '') != str(filter_branch)):
            continue
        if source and source != 'import':
            continue
        for field_name, values in (log.changes or {}).items():
            old_value, new_value = (values + ['', ''])[:2] if isinstance(values, list) else ('', values)
            client_name = f"{policy.member.first_name} {policy.member.last_name}" if policy and policy.member_id else ''
            if search_term:
                haystack = ' '.join([
                    log.membership_number or '',
                    policy.policy_number if policy else '',
                    client_name,
                    field_name,
                ]).lower()
                if search_term.lower() not in haystack:
                    continue
            entries.append({
                'date': log.import_batch.uploaded_at.strftime('%Y-%m-%d %H:%M'),
                'source': 'Import Amendment',
                'scheme': policy.scheme.name if policy and policy.scheme else '',
                'policy_number': policy.policy_number if policy else '',
                'membership_number': log.membership_number,
                'client': client_name,
                'entity': 'Policy',
                'field': field_name,
                'old_value': old_value,
                'new_value': new_value,
                'changed_by': log.import_batch.uploaded_by.username if log.import_batch.uploaded_by else 'system',
                'reference': f'AMEND-{log.import_batch_id}-ROW-{log.row_number}',
                'notes': '',
            })

    policy_ct = ContentType.objects.get_for_model(Policy)
    member_ct = ContentType.objects.get(app_label='members', model='member')
    audit_logs = AuditLog.objects.select_related('user', 'content_type').filter(
        action='update',
        content_type__in=[policy_ct, member_ct],
    ).order_by('-timestamp')
    if start_date:
        audit_logs = audit_logs.filter(timestamp__date__gte=start_date)
    if end_date:
        audit_logs = audit_logs.filter(timestamp__date__lte=end_date)

    for log in audit_logs:
        if source and source != 'audit':
            continue
        obj = log.content_object
        policy = obj if isinstance(obj, Policy) else None
        if obj is not None and not policy and hasattr(obj, 'policies'):
            policy = obj.policies.select_related('scheme', 'member').order_by('-created_at', '-id').first()
        if filter_scheme and (not policy or str(policy.scheme_id) != str(filter_scheme)):
            continue
        if filter_branch and (not policy or str(policy.scheme.branch_id if policy and policy.scheme_id else '') != str(filter_branch)):
            continue
        data = log.data or {}
        client_name = f"{policy.member.first_name} {policy.member.last_name}" if policy and policy.member_id else log.object_repr
        if search_term:
            haystack = ' '.join([
                policy.policy_number if policy else '',
                policy.membership_number if policy else '',
                client_name,
                data.get('field', ''),
                log.object_repr or '',
            ]).lower()
            if search_term.lower() not in haystack:
                continue
        entries.append({
            'date': log.timestamp.strftime('%Y-%m-%d %H:%M'),
            'source': 'Audit Update',
            'scheme': policy.scheme.name if policy and policy.scheme else '',
            'policy_number': policy.policy_number if policy else '',
            'membership_number': policy.membership_number if policy else '',
            'client': client_name,
            'entity': log.content_type.model.title(),
            'field': data.get('field', 'Record updated'),
            'old_value': data.get('old_value', ''),
            'new_value': data.get('new_value', ''),
            'changed_by': log.username,
            'reference': f'AUDIT-{log.id}',
            'notes': data.get('reason', '') or data.get('details', ''),
        })

    entries.sort(key=lambda item: item['date'], reverse=True)
    return entries


def require_all_members_access(user):
    if not can_view_all_members_report(user):
        raise PermissionDenied('You do not have access to the all members report.')


def require_payment_allocation_access(user, report_version):
    if not can_view_payment_allocation_report(user, report_version):
        raise PermissionDenied('You do not have access to this payment allocation report version.')


def require_amendments_access(user):
    if not can_view_amendments_report(user):
        raise PermissionDenied('You do not have access to the amendments report.')


def get_available_report_options(user):
    options = []

    if can_view_all_members_report(user):
        options.append({
            'key': 'all_members',
            'label': 'All Members Report',
            'description': 'Grouped member and policy listing by scheme.',
        })
    if can_view_payment_allocation_report(user, 'admin'):
        options.append({
            'key': 'payment_admin',
            'label': 'Payment Report: Admin',
            'description': 'Administrator and branch payment allocation cash-up.',
        })
    if can_view_payment_allocation_report(user, 'scheme'):
        options.append({
            'key': 'payment_scheme',
            'label': 'Payment Report: Scheme',
            'description': 'Scheme-facing payment allocation summary.',
        })
    if can_view_amendments_report(user):
        options.append({
            'key': 'amendments',
            'label': 'Amendments Report',
            'description': 'Policy changes from imports and audited updates.',
        })

    return options


def build_report_redirect(report_key, request):
    if report_key == 'all_members':
        params = {
            'branch': request.GET.get('branch', '').strip(),
            'scheme': request.GET.get('scheme', '').strip(),
            'agent': request.GET.get('agent', '').strip(),
            'status': request.GET.get('status', '').strip(),
            'search': request.GET.get('search', '').strip(),
        }
        base_url = reverse('reports:all_members_report')
    elif report_key == 'payment_admin':
        params = {
            'branch': request.GET.get('branch', '').strip(),
            'scheme': request.GET.get('scheme', '').strip(),
            'agent': request.GET.get('agent', '').strip(),
            'search': request.GET.get('search', '').strip(),
            'cover_month': request.GET.get('cover_month', '').strip(),
            'version': 'admin',
        }
        base_url = reverse('reports:payment_allocation_report')
    elif report_key == 'payment_scheme':
        params = {
            'branch': request.GET.get('branch', '').strip(),
            'scheme': request.GET.get('scheme', '').strip(),
            'agent': request.GET.get('agent', '').strip(),
            'search': request.GET.get('search', '').strip(),
            'cover_month': request.GET.get('cover_month', '').strip(),
            'version': 'scheme',
        }
        base_url = reverse('reports:payment_allocation_report')
    elif report_key == 'amendments':
        params = {
            'branch': request.GET.get('branch', '').strip(),
            'scheme': request.GET.get('scheme', '').strip(),
            'search': request.GET.get('search', '').strip(),
            'start_date': request.GET.get('start_date', '').strip(),
            'end_date': request.GET.get('end_date', '').strip(),
            'source': request.GET.get('source', '').strip(),
        }
        base_url = reverse('reports:amendments_report')
    else:
        return None

    cleaned_params = {key: value for key, value in params.items() if value not in ('', None)}
    return f"{base_url}?{urlencode(cleaned_params)}" if cleaned_params else base_url


def build_hub_url(report_key, params):
    cleaned_params = {'report': report_key}
    cleaned_params.update({key: value for key, value in params.items() if value not in ('', None)})
    query_string = urlencode(cleaned_params)
    base_url = reverse('reports:report_hub')
    return f"{base_url}?{query_string}" if query_string else base_url


@login_required
def all_members_report(request):
    require_all_members_access(request.user)
    schemes = filter_by_user_scope(Scheme.objects.select_related('branch').order_by('name'), request.user, Scheme)
    branches = get_user_branches(request.user).order_by('name')
    agents = filter_by_user_scope(Agent.objects.select_related('scheme').order_by('full_name'), request.user, Agent)
    filter_scheme = request.GET.get('scheme', '').strip()
    filter_branch = request.GET.get('branch', '').strip()
    filter_agent = request.GET.get('agent', '').strip()
    filter_status = request.GET.get('status', '').strip()
    search_term = request.GET.get('search', '').strip()

    policies = filter_by_user_scope(
        Policy.objects.select_related('member', 'scheme', 'scheme__branch', 'plan', 'underwritten_by').annotate(
            dependents_count=Count('dependents', distinct=True),
            beneficiaries_count=Count('beneficiaries', distinct=True),
        ),
        request.user,
        Policy,
    )
    if filter_branch:
        policies = policies.filter(scheme__branch_id=filter_branch)
    if filter_agent:
        policies = policies.filter(underwritten_by_id=filter_agent)
    if search_term:
        policies = policies.filter(
            Q(member__first_name__icontains=search_term)
            | Q(member__last_name__icontains=search_term)
            | Q(policy_number__icontains=search_term)
            | Q(unique_policy_number__icontains=search_term)
            | Q(membership_number__icontains=search_term)
        )
    policies = apply_policy_filters(policies, filter_scheme, filter_status).order_by('scheme__name', 'member__last_name', 'member__first_name')

    export_format = request.GET.get('export')
    if export_format:
        title, headers, rows = build_all_members_dataset(policies)
        if export_format == 'excel':
            return workbook_response(build_report_workbook(title, headers, rows), 'all-members-report.xlsx')
        if export_format == 'csv':
            return csv_response(headers, rows, 'all-members-report.csv')
        if export_format == 'pdf':
            return pdf_response(title, headers, rows, 'all-members-report.pdf')

    return render(request, 'reports/all_members_report.html', {
        'branches': branches,
        'agents': agents,
        'schemes': schemes,
        'policies': policies,
        'report_hub_url': build_hub_url('all_members', {
            'branch': filter_branch,
            'scheme': filter_scheme,
            'agent': filter_agent,
            'status': filter_status,
            'search': search_term,
        }),
        'filter_branch': filter_branch,
        'filter_agent': filter_agent,
        'filter_scheme': filter_scheme,
        'filter_status': filter_status,
        'search_term': search_term,
    })


@login_required
def payment_allocation_report(request):
    report_version = request.GET.get('version', 'admin').strip() or 'admin'
    require_payment_allocation_access(request.user, report_version)

    branches = get_user_branches(request.user).order_by('name')
    schemes = filter_by_user_scope(Scheme.objects.select_related('branch').order_by('name'), request.user, Scheme)
    agents = filter_by_user_scope(Agent.objects.select_related('scheme').order_by('full_name'), request.user, Agent)
    filter_branch = request.GET.get('branch', '').strip()
    filter_scheme = request.GET.get('scheme', '').strip()
    filter_agent = request.GET.get('agent', '').strip()
    search_term = request.GET.get('search', '').strip()
    cover_month = parse_month_param(request.GET.get('cover_month')) or timezone.now().date().replace(day=1)

    allocations = PaymentAllocation.objects.select_related(
        'payment', 'member', 'policy', 'scheme', 'branch', 'plan', 'agent'
    ).filter(
        payment__status='COMPLETED',
        allocation_status='ALLOCATED',
    )

    if not request.user.is_superuser and not user_has_role(request.user, 'Administrator'):
        allocations = allocations.filter(scheme__in=get_user_schemes(request.user))

    if filter_branch:
        allocations = allocations.filter(scheme__branch_id=filter_branch)
    if filter_scheme:
        allocations = allocations.filter(scheme_id=filter_scheme)
    if filter_agent:
        allocations = allocations.filter(agent_id=filter_agent)
    if search_term:
        allocations = allocations.filter(
            Q(member__first_name__icontains=search_term)
            | Q(member__last_name__icontains=search_term)
            | Q(policy__policy_number__icontains=search_term)
            | Q(policy__membership_number__icontains=search_term)
        )

    allocations = allocations.filter(coverage_month=cover_month).order_by(
        'scheme__name', 'payment__date', 'policy__policy_number'
    )

    wholesale_expression = ExpressionWrapper(
        F('underwriter_premium')
        + F('admin_fee')
        + F('scheme_fee')
        + F('branch_fee')
        + F('manager_fee')
        + F('cash_payout')
        + F('loyalty_programme')
        + F('other_fees'),
        output_field=DecimalField(max_digits=12, decimal_places=2),
    )
    internal_total_expression = ExpressionWrapper(
        wholesale_expression + F('agent_commission'),
        output_field=DecimalField(max_digits=12, decimal_places=2),
    )

    summary = allocations.aggregate(
        total_allocated=Sum('allocated_amount'),
        total_agent_commission=Sum('agent_commission'),
        total_wholesale=Sum(wholesale_expression),
        total_internal=Sum(internal_total_expression),
        payment_count=Count('id'),
    )

    for key in ('total_allocated', 'total_agent_commission', 'total_wholesale', 'total_internal'):
        summary[key] = summary[key] or Decimal('0.00')
    summary['payment_count'] = summary['payment_count'] or 0

    export_format = request.GET.get('export')
    if export_format:
        title, headers, rows = build_payment_allocation_dataset(allocations, cover_month, report_version)
        if export_format == 'excel':
            return workbook_response(build_report_workbook(title, headers, rows), f'payment-allocation-{cover_month.strftime("%Y-%m")}-{report_version}.xlsx')
        if export_format == 'csv':
            return csv_response(headers, rows, f'payment-allocation-{cover_month.strftime("%Y-%m")}-{report_version}.csv')
        if export_format == 'pdf':
            return pdf_response(title, headers, rows, f'payment-allocation-{cover_month.strftime("%Y-%m")}-{report_version}.pdf')

    return render(request, 'reports/payment_allocation_report.html', {
        'allocations': allocations,
        'branches': branches,
        'agents': agents,
        'schemes': schemes,
        'report_hub_url': build_hub_url(f'payment_{report_version}', {
            'branch': filter_branch,
            'scheme': filter_scheme,
            'agent': filter_agent,
            'search': search_term,
            'cover_month': cover_month.strftime('%Y-%m'),
        }),
        'filter_branch': filter_branch,
        'filter_scheme': filter_scheme,
        'filter_agent': filter_agent,
        'search_term': search_term,
        'cover_month': cover_month,
        'report_version': report_version,
        'summary': summary,
    })


@login_required
def generate_report(request):
    report_options = get_available_report_options(request.user)
    if not report_options:
        raise PermissionDenied('You do not have access to reporting.')

    selected_report = request.GET.get('report', '').strip() or report_options[0]['key']
    valid_report_keys = {option['key'] for option in report_options}
    if selected_report not in valid_report_keys:
        selected_report = report_options[0]['key']

    if request.GET.get('generate') == '1':
        redirect_url = build_report_redirect(selected_report, request)
        if redirect_url:
            return redirect(redirect_url)

    report_descriptions = {option['key']: option['description'] for option in report_options}

    return render(request, 'reports/report_hub.html', {
        'branches': get_user_branches(request.user).order_by('name'),
        'schemes': filter_by_user_scope(Scheme.objects.select_related('branch').order_by('name'), request.user, Scheme),
        'agents': filter_by_user_scope(Agent.objects.select_related('scheme').order_by('full_name'), request.user, Agent),
        'report_options': report_options,
        'selected_report': selected_report,
        'selected_report_description': report_descriptions.get(selected_report, ''),
        'filter_branch': request.GET.get('branch', '').strip(),
        'filter_scheme': request.GET.get('scheme', '').strip(),
        'filter_agent': request.GET.get('agent', '').strip(),
        'filter_status': request.GET.get('status', '').strip(),
        'search_term': request.GET.get('search', '').strip(),
        'cover_month': parse_month_param(request.GET.get('cover_month')) or timezone.now().date().replace(day=1),
        'start_date': request.GET.get('start_date', '').strip(),
        'end_date': request.GET.get('end_date', '').strip(),
        'source': request.GET.get('source', '').strip(),
    })


@login_required
def amendments_report(request):
    require_amendments_access(request.user)
    branches = get_user_branches(request.user).order_by('name')
    schemes = filter_by_user_scope(Scheme.objects.select_related('branch').order_by('name'), request.user, Scheme)
    filter_branch = request.GET.get('branch', '').strip()
    filter_scheme = request.GET.get('scheme', '').strip()
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    source = request.GET.get('source', '').strip()
    search_term = request.GET.get('search', '').strip()

    entries = collect_amendment_entries(
        filter_scheme=filter_scheme,
        filter_branch=filter_branch,
        start_date=start_date,
        end_date=end_date,
        source=source,
        search_term=search_term,
    )
    export_format = request.GET.get('export')
    if export_format:
        title, headers, rows = build_amendments_dataset(entries)
        if export_format == 'excel':
            return workbook_response(build_report_workbook(title, headers, rows), 'amendments-report.xlsx')
        if export_format == 'csv':
            return csv_response(headers, rows, 'amendments-report.csv')
        if export_format == 'pdf':
            return pdf_response(title, headers, rows, 'amendments-report.pdf')

    return render(request, 'reports/amendments_report.html', {
        'branches': branches,
        'schemes': schemes,
        'entries': entries,
        'report_hub_url': build_hub_url('amendments', {
            'branch': filter_branch,
            'scheme': filter_scheme,
            'search': search_term,
            'start_date': start_date,
            'end_date': end_date,
            'source': source,
        }),
        'filter_branch': filter_branch,
        'filter_scheme': filter_scheme,
        'start_date': start_date,
        'end_date': end_date,
        'source': source,
        'search_term': search_term,
    })
